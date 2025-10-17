from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from .forms import CustomUserCreationForm, ProductoForm, FormMensajes
from django.contrib import messages
from django.contrib.auth import authenticate, login
from .models import Producto, CanalMensaje, CanalUsuario, Canal, Ubicacion
from django.core.paginator import Paginator
from django.http import Http404, JsonResponse

from django.core.exceptions import PermissionDenied

from django.views.generic import DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic.edit import FormMixin

from django.http import JsonResponse
from django.core.exceptions import PermissionDenied
from django.views.generic.edit import FormMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Notificacion
from django.views.generic import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView
from .models import Notificacion
from django.http import Http404
from .models import Producto, Tipo

from .models import Producto, Estado, Ubicacion, Tipo
from openpyxl import Workbook

from .decorators import user_no_invitado

from django.urls import reverse

def export_to_excel_model(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="productos_exportados.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["id_producto", "id_tipo", "id_estado", "id_ubicacion", "fecha_publicacion", 'id_usuario']
    ws.append(headers)

    products = Producto.objects.all()
    for product in products:
        ws.append([
                product.id,  # Agregando el ID real
                product.tipo_id,
                product.estado_id,
                product.ubicacion_id,
                product.fecha_publicacion,
                product.usuario_id # Esto es correcto para FK directas
                ])

    wb.save(response)
    return response

def export_to_excel_estate(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="estados_exportados.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Estates"

    headers = ["id_estado", "estado_nombre"]
    ws.append(headers)

    estates = Estado.objects.all()
    for estate in estates:
        ws.append([
                estate.id,  # Agregando el ID real
                estate.estado_name,
                ])

    wb.save(response)
    return response

def export_to_excel_location(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ubicaciones_exportados.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Locations"

    headers = ["id_ubicacion", "ubicacion_nombre"]
    ws.append(headers)

    locations = Ubicacion.objects.all()
    for location in locations:
        ws.append([
                location.id,  # Agregando el ID real
                location.ubicacion_name,
                ])

    wb.save(response)
    return response

def export_to_excel_type(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="tipos_exportados.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Types"

    headers = ["id_tipo", "tipo_objeto"]
    ws.append(headers)

    types = Tipo.objects.all()
    for tipo in types:
        ws.append([
                tipo.id,  # Agregando el ID real
                tipo.tipo_objeto,
                ])

    wb.save(response)
    return response

def export_to_excel_friendly(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="productos_amigable.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos_Legibles"

    # Encabezados legibles
    headers = ["id_producto", "producto","tipo_nombre", "estado", "ubicacion_nombre", "fecha_publicacion", 'usuario']
    ws.append(headers)

    # 1. OPTIMIZACIÓN: Usar select_related para traer los datos relacionados
    productos = Producto.objects.all().select_related('tipo', 'ubicacion', 'usuario')
    
    for product in productos:
        ws.append([
            product.id,
            product.nombre,
            product.tipo.tipo_objeto if product.tipo else "N/A",  
            product.estado.estado_name if product.estado else "N/A",  
            product.ubicacion.ubicacion_name if product.ubicacion else "N/A",
            product.fecha_publicacion, 
            product.usuario.username
        ])

    wb.save(response)
    return response

def login_invitado(request):
    user = authenticate(username='invitado', password='pass_invitado')
    if user is not None:
        login(request, user)
        messages.info(request, "Has ingresado como invitado.")
        return redirect('home')
    else:
        messages.error(request, "No se pudo iniciar sesión como invitado.")
        return redirect('login')


class Inbox(View):
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.username == 'invitado':
            # Si es invitado, redirige a una página de error o a home
            messages.error(request, "Los usuarios invitados no pueden acceder a esta función.")
            return redirect(reverse('home')) # Asume que 'home' es la URL de tu página principal
        
        # Llama al dispatch del padre (FormMixin) para continuar con el procesamiento normal
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        # Marcar notificaciones como leídas al acceder al inbox
        Notificacion.objects.filter(usuario=request.user, leido=False).update(leido=True)

        inbox = Canal.objects.filter(canalusuario__usuario__in=[request.user.id])
        context = {
            "inbox": inbox,
            "notificaciones_no_leidas": Notificacion.objects.filter(usuario=request.user, leido=False).count(),
        }

        return render(request, 'mensajes/inbox.html', context)


class CanalFormMixin(FormMixin):
    form_class = FormMensajes

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.username == 'invitado':
            # Si es invitado, redirige a una página de error o a home
            messages.error(request, "Los usuarios invitados no pueden acceder a esta función.")
            return redirect(reverse('home')) # Asume que 'home' es la URL de tu página principal
        
        # Llama al dispatch del padre (FormMixin) para continuar con el procesamiento normal
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return self.request.path

    def post(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            raise PermissionDenied

        form = self.get_form()

        if form.is_valid():
            canal = self.get_object()
            usuario = self.request.user
            mensaje = form.cleaned_data.get("mensaje")
            canal_obj = CanalMensaje.objects.create(canal=canal, usuario=usuario, texto=mensaje)

            # Crear la notificación
            for miembro in canal.usuarios.exclude(id=usuario.id):  # Notificar a otros usuarios en el canal
                Notificacion.objects.create(usuario=miembro, mensaje=f"{usuario.username} te ha enviado un mensaje.")

            # Verificar si la solicitud es AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'mensaje': canal_obj.texto,
                    'username': canal_obj.usuario.username
                }, status=201)
            
            # Si no es una solicitud AJAX, continuar con el flujo normal
            return super().form_valid(form)
        
        return super().form_invalid(form)


class CanalDetailView(LoginRequiredMixin, CanalFormMixin, DetailView):
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.username == 'invitado':
            # Si es invitado, redirige a una página de error o a home
            messages.error(request, "Los usuarios invitados no pueden acceder a esta función.")
            return redirect(reverse('home')) # Asume que 'home' es la URL de tu página principal
        
        # Llama al dispatch del padre (FormMixin) para continuar con el procesamiento normal
        return super().dispatch(request, *args, **kwargs)

    template_name = 'mensajes/detail.html'
    queryset = Canal.objects.all()

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)

        canal = context['object']

        # Verificar si el usuario actual es miembro del canal
        if self.request.user not in canal.usuarios.all():
            raise PermissionDenied("No tienes permiso para ver este canal.")

        context['canal_miembro'] = self.request.user in canal.usuarios.all()
        return context

    def get_object(self, *args, **kwargs):
        username = self.kwargs.get("username")
        mi_username = self.request.user.username

        # Si el canal es de este usuario o pertenece a ambos, entonces lo mostramos
        canal, created = Canal.objects.obtener_o_crear_canal_ms(mi_username, username)

        if not canal or self.request.user not in canal.usuarios.all():
            raise Http404("Este canal no existe o no tienes acceso.")

        return canal

class DetailMs(LoginRequiredMixin, CanalFormMixin, DetailView):
    template_name = 'mensajes/detail.html'

    def get_object(self, *args, **kwargs):
        username = self.kwargs.get("username")
        mi_username = self.request.user.username
        canal, _ = Canal.objects.obtener_o_crear_canal_ms(mi_username, username)

        if username == mi_username:
            mi_canal, _ = Canal.objects.obtener_o_crear_canal_usuario_actual(self.request.user)
            return mi_canal

        if canal is None:
            raise Http404

        return canal

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        canal = self.get_object()
        context['mensaje_canal'] = canal.canalmensaje_set.all().order_by('tiempo')
        return context


@login_required
def home(request):
    query = request.GET.get('query', '')
    tipo_id = request.GET.get('tipo', '')
    ubicacion_id = request.GET.get('ubicacion', '')
    productos = Producto.objects.all()

    # Contar notificaciones no leídas
    notificaciones_no_leidas = Notificacion.objects.filter(usuario=request.user, leido=False).count()

    # Filtrar productos si hay una consulta
    if query:
        productos = productos.filter(nombre__istartswith=query)
    if tipo_id:
        productos = productos.filter(tipo_id=tipo_id)
    if ubicacion_id:
        productos = productos.filter(ubicacion_id=ubicacion_id)

    tipos = Tipo.objects.all()
    ubicaciones = Ubicacion.objects.all()  # Para el filtro de región

    page = request.GET.get('page', 1)

    try:
        paginator = Paginator(productos, 6)
        productos = paginator.page(page)
    except:
        raise Http404

    return render(request, 'app/index.html', {
        'entity': productos,
        'paginator': paginator,
        'query': query,
        'tipos': tipos,
        'tipo_seleccionado': tipo_id,
        'notificaciones_no_leidas': notificaciones_no_leidas,
        'ubicaciones': ubicaciones,
        'ubicacion_seleccionada': ubicacion_id,
    })

def register(request):
    return render(request, 'app/register.html');

@user_no_invitado
@login_required
def producto(request, id):
    entidad = get_object_or_404(Producto, id=id)
    data = {'entidad': entidad, 'latitud': entidad.latitud, 'longitud': entidad.longitud}
    return render(request, 'app/producto.html', data)

def registro(request):
    data = {
        'form': CustomUserCreationForm()
    }

    if request.method == 'POST':
        formulario = CustomUserCreationForm(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            username = formulario.cleaned_data["username"]
            password = formulario.cleaned_data["password1"]
            user = authenticate(username=username, password=password)

            if user is not None:
                login(request, user)
                messages.success(request, "Te has registrado correctamente")
                return redirect(to="home")
            else:
                messages.error(request, "Error en la autenticacion. Por favor, intentalo de nuevo.") 
        data['form'] = formulario

    return render(request, 'registration/register.html', data);

@user_no_invitado
@login_required
def agregar_producto(request):
    data = {
        'form': ProductoForm(user=request.user)
    }

    if request.method == 'POST':
        formulario = ProductoForm(data=request.POST, files=request.FILES, user=request.user)

        if formulario.is_valid():
            formulario.instance.usuario = request.user

            # Obtener latitud y longitud
            latitude = request.POST.get('latitude', '').strip()
            longitude = request.POST.get('longitude', '').strip()

            if latitude and longitude:
                try:
                    formulario.instance.latitud = float(latitude)
                    formulario.instance.longitud = float(longitude)
                except ValueError:
                    messages.error(request, "Latitud y longitud deben ser números válidos.")
                    data['form'] = formulario
                    return render(request, 'app/producto/agregar.html', data)
            else:
                formulario.instance.latitud = None
                formulario.instance.longitud = None

            formulario.save()
            messages.success(request, "Agregado correctamente")

            # ⬇️ Cambio aquí: redirigimos con la página actual
            page = request.GET.get("page", 1)
            return redirect(f"{reverse('listar-productos')}?page={page}")

        else:
            data["form"] = formulario

    return render(request, 'app/producto/agregar.html', data)

@user_no_invitado
@login_required
def listar_productos(request):
    # Filtrar los productos según el usuario
    if request.user.is_superuser:
        productos = Producto.objects.all()
    else:
        productos = Producto.objects.filter(usuario=request.user)

    query = request.GET.get('query', '')
    tipo_id = request.GET.get('tipo', '')
    ubicacion_id = request.GET.get('ubicacion', '')

    if query:
        productos = productos.filter(nombre__istartswith=query)
    if tipo_id:
        productos = productos.filter(tipo_id=tipo_id)
    if ubicacion_id:
        productos = productos.filter(ubicacion_id=ubicacion_id)

    tipos = Tipo.objects.all()
    ubicaciones = Ubicacion.objects.all()

    page = request.GET.get('page', 1)

    try:
        paginator = Paginator(productos, 6)
        productos = paginator.page(page)
    except:
        raise Http404

    context = {
        'entity': productos,
        'query': query,
        'paginator': paginator,
        'tipos': tipos,
        'tipo_seleccionado': tipo_id,
        'ubicaciones': ubicaciones,
        'ubicacion_seleccionada': ubicacion_id,
    }
    return render(request, 'app/producto/listar.html', context)


from django.urls import reverse

@user_no_invitado
@login_required
def modificar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    # Verificar si el usuario actual es el dueño del producto o es un superusuario
    if request.user != producto.usuario and not request.user.is_superuser:
        raise PermissionDenied("No tienes permiso para modificar este producto.")

    data = {
        'form': ProductoForm(instance=producto),
        'entidad': producto  # Para pasar los valores de latitud y longitud al template
    }

    if request.method == 'POST':
        formulario = ProductoForm(data=request.POST, instance=producto, files=request.FILES)
        if formulario.is_valid():
            # Obtener y limpiar latitud y longitud
            latitude = request.POST.get('latitude', '').strip()
            longitude = request.POST.get('longitude', '').strip()

            try:
                if latitude:
                    producto.latitud = float(latitude)
                if longitude:
                    producto.longitud = float(longitude)
            except ValueError:
                messages.error(request, "Latitud y longitud deben ser números válidos.")
                data['form'] = formulario
                return render(request, 'app/producto/modificar.html', data)

            formulario.save()
            messages.success(request, "Modificado correctamente")

            # ⬇️ Mantener la página actual
            page = request.GET.get("page", 1)
            return redirect(f"{reverse('listar-productos')}?page={page}")
        else:
            data["form"] = formulario

    # ⬇️ Este return asegura que siempre devolvemos un HttpResponse
    return render(request, 'app/producto/modificar.html', data)


@user_no_invitado
@login_required
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.user != producto.usuario and not request.user.is_superuser:
        raise PermissionDenied("No tienes permiso para eliminar este producto.")
    
    producto.delete()
    page = request.GET.get("page", 1)  # capturamos la página actual
    messages.success(request, "Producto eliminado correctamente")
    return redirect(f"{reverse('listar-productos')}?page={page}")

@user_no_invitado
@login_required
def mensajes_privados(request, username, *args, **kwargs):
    if not request.user.is_authenticated:
        return HttpResponse("Prohibido")

    mi_username = request.user.username
    canal, created = Canal.objects.obtener_o_crear_canal_ms(mi_username, username)

    if created:
        print("Si, fue creado")

    # Ordenar los mensajes aquí
    mensaje_canal = canal.canalmensaje_set.all().order_by('tiempo')

    context = {
        'canal': canal,
        'mensaje_canal': mensaje_canal,
        'usuarios_canal': canal.canalusuario_set.all(),
    }

    return render(request, 'mensajes/detail.html', context)



# --- PayPal / Suscripciones ---
import json
from decimal import Decimal
from django.conf import settings
from django.http import JsonResponse
from django.utils import timezone

from paypalcheckoutsdk.orders import OrdersCreateRequest, OrdersCaptureRequest

from .models import SubscriptionPayment, Profile  # <-- añadimos Profile

@user_no_invitado
@login_required
def paypal_create_order(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    # Cliente PayPal desde tu helper
    from .paypal_client import get_paypal_client
    client = get_paypal_client()

    create_request = OrdersCreateRequest()
    create_request.prefer("return=representation")
    create_request.request_body({
        "intent": "CAPTURE",
        "purchase_units": [{
            "amount": {
                "currency_code": settings.SUBSCRIPTION_CURRENCY,
                "value": str(settings.SUBSCRIPTION_PRICE.quantize(Decimal('1.00'))),
            },
            "description": settings.SUBSCRIPTION_DESCRIPTION if hasattr(settings, "SUBSCRIPTION_DESCRIPTION") else "Suscripción Premium"
        }]
    })

    response = client.execute(create_request)
    order = response.result
    return JsonResponse({"id": order.id})

@user_no_invitado
@login_required
def paypal_capture_order(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        body = json.loads(request.body.decode("utf-8"))
    except Exception:
        body = {}

    order_id = body.get("orderID")
    if not order_id:
        return JsonResponse({"error": "Missing orderID"}, status=400)

    from .paypal_client import get_paypal_client
    client = get_paypal_client()

    capture_request = OrdersCaptureRequest(order_id)
    capture_request.request_body({})

    try:
        response = client.execute(capture_request)
        result = response.result  # objeto Python, NO tiene .json()

        status = getattr(result, "status", "")

        # Tomamos datos útiles del objeto para guardar algo legible
        purchase_units = getattr(result, "purchase_units", []) or []
        units_slim = []
        for u in purchase_units:
            amt = getattr(u, "amount", None)
            units_slim.append({
                "reference_id": getattr(u, "reference_id", None),
                "amount": {
                    "currency_code": getattr(amt, "currency_code", None) if amt else None,
                    "value": getattr(amt, "value", None) if amt else None,
                }
            })

        raw_payload = {
            "id": getattr(result, "id", None),
            "status": status,
            "payer_email": getattr(getattr(result, "payer", None), "email_address", None),
            "purchase_units": units_slim
        }

        # Guarda el pago (asumo que raw es JSONField o TextField)
        SubscriptionPayment.objects.create(
            user=request.user,
            order_id=order_id,
            status=status,
            amount=settings.SUBSCRIPTION_PRICE,
            currency=settings.SUBSCRIPTION_CURRENCY,
            raw=raw_payload  # <<--- ya no usamos result.json()
        )

        # Asegurar que el usuario tenga Profile
        profile, _ = Profile.objects.get_or_create(user=request.user)

        if status == "COMPLETED":
            # Marcar Premium al usuario
            profile.is_premium = True
            profile.premium_since = timezone.now()
            profile.save()
            return JsonResponse({"status": "COMPLETED"})
        else:
            return JsonResponse({"status": status})

    except Exception as e:
        # Devuelve mensaje para verlo en la consola del navegador
        return JsonResponse({"error": str(e)}, status=500)

@user_no_invitado
@login_required
def subscription_success(request):
    return render(request, "app/subscription_success.html")

def export_to_excel(request):
    # Esta vista puede redirigir o mostrar una página con enlaces a las diferentes exportaciones
    return render(request, 'app/exportar.html')